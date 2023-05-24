import math

import openpyxl
import os
import re

Structural_Element = "Laterally Unsupported Beam"

folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads', "Advance Steel Design")
steel_table = os.path.join(folder, 'indian-steel-table.xlsx')
design_sheet= os.path.join(folder, 'Laterally Unsupported Beam.xlsm')
result_sheet = os.path.join(folder, f'{Structural_Element}.xlsx')
wb1 = openpyxl.load_workbook(design_sheet)
ws1 = wb1["Laterally Unsupported Beams"]

wb = openpyxl.load_workbook(steel_table)

wb2 = openpyxl.Workbook()
ws1._parent = wb2
wb2._add_sheet(ws1)
ws2 = wb2["Laterally Unsupported Beams"]




Beams_Sections = ["ISJB", "ISLB", "ISMB", "ISWB", "ISHB"]

def column_return(name, ws):
    max_column = ws.max_column

    for i in range(1, max_column + 1):
        # try:
        texts = ws.cell(row=3, column=i).value
        if texts  == None:
            texts = ws.cell(row=1, column=i).value


        try:
            texts = texts.split(" ")


            for text in texts:
                if text == name:
                    return i
        except:
            print(texts)





M_design = 78
j = 2

for beam in Beams_Sections:
    ws = wb[beam]
    max_row = ws.max_row
    Lt = ws1.cell(row = 4, column = 6).value
    E = ws1.cell(row = 5, column = 6).value
    Fy = ws1.cell(row=6, column=6).value
    Beta = ws1.cell(row=7, column=6).value
    AlphaLt = ws1.cell(row=8, column=6).value
    GammaMo = ws1.cell(row=9, column=6).value


    for i in range(5, max_row):
        Ixx = ws.cell(row=i, column=8).value           #cm
        Iyy = ws.cell(row = i, column = 9).value       #cm
        rxx = ws.cell(row=i, column=10).value          #cm
        ryy = ws.cell(row = i, column = 11).value      #cm
        h = ws.cell(row=i, column=4).value             #mm
        tf =  ws.cell(row=i, column=6).value           #mm
        Zxx =  ws.cell(row=i, column=12).value
        Zyy =  ws.cell(row=i, column=13).value


        if (Ixx != None) and (Ixx !=0):
            # print(Ixx)
            Lt = Lt
            Ixx = Ixx / 100000000
            Iyy = Iyy / 100000000
            rxx = rxx / 100
            ryy = ryy / 100
            tf = tf / 1000
            h = h / 1000
            Zxx = Zxx / 1000000
            Zyy = Zyy / 1000000




            hf = h - tf
            Ze = Zxx
            Zp = Zxx + Zyy
            Mcr =  ((((math.pi) ** 2) * E * Iyy * hf ) / (2 * (Lt**2))) * ((1 + ((((Lt / ryy) / (hf / tf))**2)/20)) ** 0.5)
            Fcrb = ((1.1 * ((math.pi) ** 2) * E) / ((Lt / ryy) ** 2)) * ((1 + ((((Lt / ryy) / (hf / tf))**2)/20)) ** 0.5)


            LambdaLt1 = (Beta * Zp * Fy / Mcr ) ** 0.5
            LambdaLt2 = (1.2 * Ze * Fy / Mcr ) ** 0.5
            LambdaLt = ( Fy / Fcrb ) ** 0.5

            if LambdaLt1 < LambdaLt2:
                LambdaLt = LambdaLt1

            PhiLt = 0.5*(1 + AlphaLt *(LambdaLt - 0.2) + (LambdaLt ** 2))

            XhiLt = 1/(PhiLt + (((PhiLt ** 2) - (LambdaLt **2))**0.5))

            Fbd = XhiLt * Fy / GammaMo

            Md = Beta * Zp * Fbd

            ws2.cell(row=20, column=j).value = ws.cell(row=i, column=1).value
            ws2.cell(row=21, column=j).value = Mcr
            ws2.cell(row=22, column=j).value = Fcrb
            ws2.cell(row=23, column=j).value = LambdaLt
            ws2.cell(row=24, column=j).value = PhiLt
            ws2.cell(row=25, column=j).value = XhiLt
            ws2.cell(row=26, column=j).value = Fbd
            ws2.cell(row=27, column=j).value = Md
            j +=1






wb2.save(result_sheet)
