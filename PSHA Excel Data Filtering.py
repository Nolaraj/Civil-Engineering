import openpyxl
from tkinter import Tk
import pdf2image
import pyautogui, time
import os
from tkinter import Tk, filedialog
from PIL import Image
from time import sleep
import os
import pdf2image

def path_finder():
    root = Tk() # pointing root to Tk() to use it as Tk() in program.
    root.withdraw() # Hides small tkinter window.
    root.title()
    root.attributes('-topmost', True) # Opened windows will be active. above all windows despite of selection.
    # root.attributes()
    open_file = filedialog.askopenfilename(title = "Select file",filetypes = (("Excel files", ".xlsx .xls"),))  # Returns opened path as str
    return open_file
file_path = path_finder()
file_dir = open_file = filedialog.askdirectory(title = "Select Directory")#,filetypes = (("Excel files", ".xlsx .xls"),))
os.chdir(file_dir)
# filename2 = filedialog.askopenfilename(title = "Select file",filetypes = (("Excel files", ".xlsx .xls"),))
workbook = openpyxl.load_workbook(file_path, data_only=True)
worksheet = workbook.active

title = []
for i in range(1,8):
    title.append(worksheet.cell(row=1, column=i).value)

last_row = worksheet.max_row
year = 2022
row = 1
# for i in range (2,last_row):
#     print(year)
#     if year == worksheet.cell(row=i, column=7).value:
#         data = []
#         for j in range(1, 8):
#             data.append(worksheet.cell(row=i, column=j).value)
#         for k in range(0,7):
#             worksheet.cell(row=row, column=10+k).value = data[k]
#         row += 1
#
#     else:
#         year = worksheet.cell(row=i, column=7).value
#         data = []
#         for j in range(1, 8):
#             data.append(worksheet.cell(row=i, column=j).value)
#         row += 3
#         for k in range(0,7):
#             worksheet.cell(row=row, column=10+k).value = title[k]
#             worksheet.cell(row=row+1, column=10+k).value = data[k]
#         row += 2

for i in range (2,last_row):
    if (year == worksheet.cell(row=i, column=7).value) or (int(year) - 10 < int(worksheet.cell(row=i, column=7).value)) :
        data = []
        for j in range(1, 8):
            data.append(worksheet.cell(row=i, column=j).value)
        for k in range(0,7):
            worksheet.cell(row=row, column=20+k).value = data[k]
        row += 1

    else:
        year = worksheet.cell(row=i, column=7).value
        data = []
        for j in range(1, 8):
            data.append(worksheet.cell(row=i, column=j).value)
        row += 3
        for k in range(0,7):
            worksheet.cell(row=row, column=20+k).value = title[k]
            worksheet.cell(row=row+1, column=20+k).value = data[k]
        row += 2
