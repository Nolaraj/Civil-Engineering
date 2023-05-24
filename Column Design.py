import math

import openpyxl
import os
import re

Structural_Element = "Column Design"

folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads', "Advance Steel Design")
steel_table = os.path.join(folder, 'indian-steel-table.xlsx')
design_sheet= os.path.join(folder, 'Steel Design.xlsm')
result_sheet = os.path.join(folder, f'{Structural_Element}.xlsx')
wb1 = openpyxl.load_workbook(design_sheet)
ws1 = wb1[Structural_Element]

wb = openpyxl.load_workbook(steel_table)
ws_Fcd = wb["Fcd - Buckling Class C"]

wb2 = openpyxl.Workbook()
ws1._parent = wb2
wb2._add_sheet(ws1)
ws2 = wb2[Structural_Element]




Columns_Sections = ["ISJC", "ISLC", "ISMC"]

def column_return(name, ws):
    max_column = ws.max_column

    for i in range(1, max_column + 1):
        # try:
        texts = ws.cell(row=3, column=i).value
        if texts  == None:
            texts = ws.cell(row=1, column=i).value


        try:
            texts = texts.split(" ")


            for text in texts:
                if text == name:
                    return i
        except:
            print(texts)






j = 2

for column in Columns_Sections:
    ws = wb[column]
    max_row = ws.max_row
    L = ws1.cell(row = 4, column = 6).value
    E = ws1.cell(row = 5, column = 6).value
    P = ws1.cell(row = 6, column = 6).value
    Fu = ws1.cell(row = 7, column = 6).value
    Fy = ws1.cell(row = 8, column = 6).value
    k = ws1.cell(row = 9, column = 6).value
    Design_Stress = ws1.cell(row = 10, column = 6).value
    Channel_Orientation = bool(ws1.cell(row = 11, column = 6).value)
    N = bool(ws1.cell(row = 12, column = 6).value)


    Bolt_Grade = ws1.cell(row = 16, column = 6).value
    Fub = ws1.cell(row = 17, column = 6).value
    GammaMo = ws1.cell(row= 18, column=6).value
    GammaMb = ws1.cell(row= 19, column=6).value
    Bolt_Dia = ws1.cell(row = 21, column = 6).value




    for i in range(5, max_row):
        A = ws.cell(row=i, column=3).value
        Izz = ws.cell(row=i, column=9).value           #cm Ixx
        Iyy = ws.cell(row = i, column = 10).value       #cm
        rz = ws.cell(row=i, column=11).value          #cm rx
        ry = ws.cell(row = i, column = 12).value      #cm
        h = ws.cell(row=i, column=4).value             #mm
        Cyy =  ws.cell(row=i, column=8).value           #mm
        bf =  ws.cell(row=i, column=5).value
        Bolt_Dia = Bolt_Dia/1000


        if (A != None) and (A !=0):
            # print(Ixx)
            L = L
            A = A / 10000
            Izz = Izz / 100000000
            Iyy = Iyy / 100000000
            rz = rz / 100
            ry = ry / 100
            h = h / 1000
            Cyy = Cyy / 100
            bf = bf / 1000

            Area_req =  P/Design_Stress
            Area_Provided = 2 * A

            Se = 1.1 * (k * L / rz)

            Fcd = 136.27
            for d in range(5, 29):
                v1 = (math.floor(Se/10) * 10)
                v2 = math.ceil(Se/10) * 10

                v1_value1 = int(ws_Fcd.cell(row = d, column = 2).value)


                v1_value = int(ws_Fcd.cell(row = d, column = 3).value)
                v2_value = int(ws_Fcd.cell(row = d + 1, column = 3).value)

                if v1_value1 == v1:
                    Fcd = v1_value + ((v2_value - v1_value) / 10)*(Se - v1)
                    break

            Pd = Fcd * Area_Provided

            S = 2*((((Izz-Iyy)/A)**0.5) - Cyy)      #m
            S = S*1000                              #mm
            S = math.ceil(S/100) * 100
            S = S/1000                              #m



            C1 = 0.7 * Se * ry
            C2 = 50 * ry
            C0 = min(C1, C2)
            C = C0 * 1000                           #mm
            C = math.floor(C/100) * 100
            C = C0 / 1000                           #m

            # print(C)
            if Bolt_Dia > 18:
                Dh = Bolt_Dia*1000 + 2
            else:
                Dh = Bolt_Dia*1000 + 1.5

            Ed = 1.5 * Dh/1000

            #End Battens
            D_prime = S + 2*Cyy
            while D_prime < 2*bf:
                S = S + 10/1000
                D_prime = S + 2 * Cyy
            D_Overall_E = D_prime + 2 * Ed

            Tb = math.ceil((S + 2 * bf)*1000 / 50)
            if Tb < 6:
                Tb = 6
            Tb = Tb / 1000

            #Intermediate Battens
            D_prime_I = D_prime * 3 / 4
            if D_prime_I < 2*bf:
                D_prime_I = 2*bf
            D_Overall_I = D_prime_I + 2*Ed

            Tb= Tb

            # Dimensions

            #Design Forces
            Vt = 0.025 * P
            Vb = Vt * C / (2 * (200 + 2*50))
            M = Vt*C / (2 * N)

            Shear_Stress = Vb / (D_Overall_I * Tb)

            Shear_check = True
            if Shear_Stress < Fy / ((3 ** 0.5) * 1.1):
                Shear_check = False

                Shear_Stress = Vb / (D_Overall_I * Tb)

            Bending_Stress = 6 * M / (D_Overall_I * D_Overall_I * Tb)
            Moment_check = True

            if Shear_Stress < Fy /  1.1:
                Moment_check = False

            #Connections Design
            Anb = (0.78 * math.pi * Bolt_Dia ** 2)/4
            Bolt_strength_1 = Anb * Fub /((3 ** 0.5) * GammaMo)
            Bolt_strength_2 = 2.5 * 1 * Bolt_Dia * Tb * Fu / GammaMb

            Bolt_Strength = max(Bolt_strength_1, Bolt_strength_2)

            Bolts_Number = math.ceil(Vb / Bolt_Strength)
            if Bolts_Number < 3:
                Bolts_Number = 3

            Pitch = (D_Overall_I - 2*Ed) * 0.5



            Shear_Force = Vb / Bolts_Number
            sum_r2 = 0

            BN = Bolts_Number
            print("Loop Started", BN)


            I_N = 1
            while BN > 0:
                sum_r2 = (Bolts_Number - 1) *( (I_N * Pitch)**2)
                BN = BN - 2
                I_N = I_N + 1

            print("Loop Ended")

            Moment_Force = M * Pitch/sum_r2

            Resultant = (Shear_Force + Moment_Force)**0.5



            ws2.cell(row=24, column=j).value = ws.cell(row=i, column=1).value
            ws2.cell(row=25, column=j).value = Area_req
            ws2.cell(row=26, column=j).value = Area_Provided
            ws2.cell(row=27, column=j).value = Se
            ws2.cell(row=28, column=j).value = Fcd
            ws2.cell(row=29, column=j).value = Pd
            ws2.cell(row=30, column=j).value = S
            ws2.cell(row=31, column=j).value = C
            ws2.cell(row=32, column=j).value = Dh
            ws2.cell(row=33, column=j).value = Ed
            ws2.cell(row=35, column=j).value = D_Overall_E
            ws2.cell(row=36, column=j).value = Tb
            ws2.cell(row=38, column=j).value = D_Overall_I
            ws2.cell(row=39, column=j).value = Tb
            ws2.cell(row=41, column=j).value = Vb
            ws2.cell(row=42, column=j).value = Vt
            ws2.cell(row=43, column=j).value = M
            ws2.cell(row=44, column=j).value = Shear_Stress
            ws2.cell(row=45, column=j).value = Bending_Stress
            ws2.cell(row=47, column=j).value = Bolt_Strength
            ws2.cell(row=48, column=j).value = Bolts_Number
            ws2.cell(row=49, column=j).value = Pitch
            ws2.cell(row=50, column=j).value = Shear_Force
            ws2.cell(row=51, column=j).value = Moment_Force
            ws2.cell(row=52, column=j).value = Resultant



            j +=1






wb2.save(result_sheet)
