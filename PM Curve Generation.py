import math
import os
from math import *
from openpyxl import *
import openpyxl
import matplotlib.pyplot as plt
import numpy as np

'''Input Region, Values are provided in m, KN'''
##Characteristic Strengths of concrete and Steel in MPa
fck = 20
fy = 500
#For Section
#Outer diameters of hollow circular sections in m
outer_dias = [2, 3 , 4, 5]   #Outer Diameter can be added in any numbers
#Thickess of hollow section (Inner dia  =Outer dia - thickness), If hollow  = False, The section is solid
t = 0.75
hollow = True
str_increment = 0.25

#For Reinforcements (mm)
#Outer Layer data
#Rebar diameter and number of rebar layers along other part of section
d_outer = 32
number_outer = 1
#Effective cover for section (mm)
eff_cover = 100
inner_layer = True

#Inner Layer Data
#Inner layer parameters (rebar diameter and number of layers)
d_inner = d_outer
number_inner = 4


#Strain Data beyound 0.002[end1, end2]
end_epsilon = [[0.002, 0.00225, 0.0025, 0.00275, 0.003, 0.00325, 0.0035, 0.0035, 0.0035, 0.0035, 0.0035, 0.0035,
               0.0035, 0.0035, 0.0035, 0.0035] , [0.002, 0.0017, 0.0013, 0.001, 0.0007, 0.0003, 0, -0.0025,
                                                  -0.005, -0.0075, -0.01, -0.0125, -0.015, -0.0175, -0.02, -0.0225]]




'''Processing Section'''
strains = [i/100000 for i in range(0, 201, 5)]
total_rows = 1 + 1 + len(strains) + 1 + 20

for number in range(len(outer_dias)):
    D = outer_dias[number]
    d = d_outer
    # Reinforcement coordinates from leftmost top corner of inscribed circle inside square
    outer_rein_d = (D - 2 * (eff_cover / 1000))
    outer_rein_r = outer_rein_d / 2
    inner_rein_d = ((D - 2 * t) + 2 * (eff_cover / 1000))
    inner_rein_r = inner_rein_d / 2
    outer_ring = pi * outer_rein_d
    inner_ring = pi * ((D - 2 * t) + 2 * (eff_cover / 1000))
    outer_rein_spacing = outer_ring / number_outer
    inner_rein_spacing = inner_ring / number_inner
    alpha_a = outer_rein_spacing * 360 / outer_ring
    alpha_b = inner_rein_spacing * 360 / inner_ring

    # Outer layer append
    rein_position = []
    x = D / 2
    y = eff_cover / 1000

    for i in range(0, number_outer):
        alpha_1 = alpha_a * i
        if i == 0:
            rein_position.append([x, y])
        else:
            dx = outer_rein_r * sin(radians(alpha_1))
            dy = outer_rein_r * (1 - cos(radians(alpha_1)))

            rein_position.append([x + dx, y + dy])

    x = D / 2
    y = (t - eff_cover / 1000)
    for i in range(0, number_inner):
        alpha_2 = alpha_b * i
        if i == 0:
            rein_position.append([x, y])
        else:
            dx = inner_rein_r * sin(radians(alpha_2))
            dy = inner_rein_r * (1 - cos(radians(alpha_2)))
            rein_position.append([x + dx, y + dy])

    R = D / 2
    iteration_no = int(D / str_increment)
    ap = 0.00
    dx = str_increment
    inner_r = R - t
    inner_d = 2 * inner_r

    sn = 1
    total_steel = 0.000


    def PM_generator(strain_1end, strain_2end):
        global total_steel
        global sn
        inner_area_psegment = 0.000
        area_psegment = 0.000
        sum_P = 0.000
        sum_M = 0.00
        inner_area = 0.00

        for i in range(0, iteration_no):
            a = float(str_increment * (i + 1))
            # Area of Outermost concrete strip
            theta = 2 * math.degrees(acos((R - a) / R))
            area_asegment = (pi * D * D * theta / (4 * 360)) - (
                    R * R * sin(radians(theta / 2)) * cos(radians(theta / 2)))
            area_strip = area_asegment - area_psegment
            area_psegment = area_asegment
            area_strip = area_strip * 1000000
            # Inner Circular section
            if ((a - dx) < t) and (a > t):
                a1 = a - t
                theta = 2 * math.degrees(acos((inner_r - a1) / inner_r))
                inner_area_asegment = (pi * inner_d * inner_d * theta / (4 * 360)) - (
                        inner_r * inner_r * sin(radians(theta / 2)) * cos(radians(theta / 2)))
                inner_area_strip = inner_area_asegment - inner_area_psegment
                inner_area_psegment = inner_area_asegment
            elif ((a - dx) >= t) and (a < (t + inner_d)):
                a1 = a - t
                theta = 2 * math.degrees(acos((inner_r - a1) / inner_r))
                inner_area_asegment = (pi * inner_d * inner_d * theta / (4 * 360)) - (
                        inner_r * inner_r * sin(radians(theta / 2)) * cos(radians(theta / 2)))
                inner_area_strip = inner_area_asegment - inner_area_psegment
                inner_area_psegment = inner_area_asegment
            elif ((a - dx) < (t + inner_d)) and (a >= (t + inner_d)):
                a1 = (t + inner_d)
                inner_area_asegment = pi * inner_d * inner_d / 4
                inner_area_strip = inner_area_asegment - inner_area_psegment
                inner_area_psegment = inner_area_asegment
            else:
                inner_area_asegment = 0
                inner_area_strip = 0
                inner_area_psegment = 0
            inner_area_strip = inner_area_strip * 1000000

            inner_area = inner_area_strip + inner_area
            # Area of the steel under considered strip
            area_steel = 0.00
            for item in rein_position:
                lr = item[1]
                if ((a - dx) <= (lr - d / (2 * 1000)) and (a >= (lr + d / (2 * 1000)))):
                    area_steel = area_steel + pi * d * d / 4
                elif (a <= (lr + d / (2 * 1000)) and (a >= (lr - d / (2 * 1000)))):
                    if a > lr:
                        r = d / 2
                        b = sqrt(r * r - (a - lr) * (a - lr))
                        angle = 2 * math.degrees(asin(b / r))
                        area_steel = area_steel + (pi * d * d * (360 - angle) / (4 * 360)) + (b * (a - lr))
                    elif a < lr:
                        r = d / 2
                        b = sqrt(r * r - (-a + lr) * (-a + lr))
                        angle = 2 * math.degrees(asin(b / r))
                        area_steel = area_steel + (pi * d * d * (angle) / (4 * 360)) - (b * (lr - a))
                    else:
                        area_steel = area_steel + (pi * d * d / 8)
                elif ((a - dx) <= (lr + d / (2 * 1000)) and ((a - dx) >= (lr - d / (2 * 1000)))):
                    if ap > lr:
                        r = d / 2
                        b = sqrt(r * r - (ap - lr) * (ap - lr))
                        angle = 2 * math.degrees(asin(b / r))
                        area_steel = area_steel + (pi * d * d * (angle) / (4 * 360)) - (b * (ap - lr))
                    elif ap < lr:
                        r = d / 2
                        b = sqrt(r * r - (-ap + lr) * (-ap + lr))
                        angle = 2 * math.degrees(asin(b / r))
                        area_steel = area_steel + (pi * d * d * (360 - angle) / (4 * 360)) + (b * (lr - ap))
                    else:
                        area_steel = area_steel + (pi * d * d / 8)
                else:
                    pass
                ap = a
            total_steel = total_steel + area_steel
            # *------------------------------------------------------------------------
            # Net concrete Area
            concrete_area = area_strip - inner_area_strip - area_steel
            # Strain on considered Strip
            x = a - dx / 2
            epsilon_a = strain_1end
            epsilon_b = strain_2end
            epsilon = epsilon_a + (x / D) * (epsilon_b - epsilon_a)
            # Stresses @ Considered strip
            if abs(epsilon) >= 0.002:
                conc_stress = 0.446 * fck * (epsilon / abs(epsilon))
            else:
                conc_stress = 0.446 * fck * epsilon * (epsilon / abs(epsilon + 0.0000000000000000000000000001)) / 0.002
            if abs(epsilon) >= 0.00217:
                steel_stress = fy * (epsilon / abs(epsilon)) / 1.15
            else:
                steel_stress = fy * epsilon / (0.00217 * 1.15)
            # Force Resisted by the strip
            P = conc_stress * concrete_area + steel_stress * area_steel
            P = P / 1000  # In KN
            sum_P = sum_P + P

            # Moment Created considering from top fiber
            M = P * (a - dx / 2)
            sum_M = sum_M + M
        # Eccentricity Determination
        e = (sum_M / sum_P) - R
        Force = sum_P
        Moment = abs(sum_P * e)

        # Excel Writing for the data extracted
        f_result = "PM Data.xlsx"
        sheet = "Result"
        main_folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Documents', 'PM Curve Generation')
        if os.path.exists(main_folder) is False:
            os.makedirs(main_folder)
        os.chdir(main_folder)
        try:
            wb = load_workbook(f_result)
            ws = wb[sheet]
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet
        title = ["S.N", "Strain top", "Strain top", "Force", "Moment"]
        ws.cell(row=3 + total_rows*number, column=1).value = f'For diameter of reinforcement as {d_outer}'
        ws.cell(row=4 + total_rows*number, column=1).value = f'PM Curve Generation for hollow circular setion ' \
                                                                  f'with outer dia of {D} and thickness of {t}'

        for i in range(len(title)):
            ws.cell(row=5  + total_rows*number, column=i + 1).value = title[i]
        ws.cell(row=5 + sn + total_rows*number, column=1).value = sn
        ws.cell(row=5 + sn + total_rows*number, column=2).value = epsilon_a
        ws.cell(row=5 + sn + total_rows*number, column=3).value = epsilon_b
        ws.cell(row=5 + sn + total_rows*number, column=4).value = Force
        ws.cell(row=5 + sn + total_rows*number, column=5).value = Moment

        wb.save(filename=f_result)
        sn = sn + 1


    # Calling PM_generator and providing data of strain
    # For less than 0.002
    for items in strains:
        strain_1end = items
        strain_2end = 0.0035 - (0.75 * strain_1end)
        PM_generator(strain_1end, strain_2end)

    for i in range(len(end_epsilon[0])):
        strain_1end = end_epsilon[0][i]
        strain_2end = end_epsilon[1][i]
        PM_generator(strain_1end, strain_2end)



#Chart Creation
f_result = "PM Data.xlsx"
sheet = "Result"
try:
    wb = load_workbook(f_result)
    ws = wb[sheet]
except:
    print("Chart Not created")

chart = openpyxl.chart.ScatterChart()
chart.title = "PM Curve For Hollow Section"
chart.style = 13
chart.x_axis.title = 'Moment (kN.m)'
chart.y_axis.title = 'Force(kN)'
# Inputting data for plotting
xvalues=None
values = None
series= None
for number in range(len(outer_dias)):
    xvalues = openpyxl.chart.Reference(ws, min_col= 5 , min_row=6 + total_rows*number, max_row=sn + 4 + total_rows*number)
    values = openpyxl.chart.Reference(ws, min_col= 4 , min_row=6  + total_rows*number, max_row=sn + 4 + total_rows*number)
    series = openpyxl.chart.Series(values, xvalues, title_from_data=False)



    chart.series.append(series)

    XMag = []
    YMag = []
    for row in range(6 + total_rows*number, sn + 4 + total_rows*number):
        XMag.append(ws.cell(column = 5, row = row).value)
        YMag.append(ws.cell(column = 4, row = row).value)
    XMag = np.array(XMag)
    YMag = np.array(YMag)
    plt.plot(XMag, YMag)
    plt.show()

ws.add_chart(chart, "H8")

wb.save(filename=f_result)

print(os.getcwd())
