from math import *

import openpyxl
import os

Structural_Element = "Howe Truss"

folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop', "Python Programming", "Report Automation Project", "Standard Files Revised 1")
Truss_file = os.path.join(folder, 'Truss Optimization', "Truss.xlsx")

wb = openpyxl.load_workbook(Truss_file)
ws = wb[Structural_Element]


L = ws.cell(row =3 , column = 6).value
H = ws.cell(row =4 , column = 6).value
Pav = ws.cell(row =5 , column = 6).value
Pah = ws.cell(row =6, column = 6).value
Pdv = ws.cell(row =7 , column = 6).value
Pdh = ws.cell(row =8 , column = 6).value
Pfv = ws.cell(row =9 , column = 6).value
Pfh = ws.cell(row =10 , column = 6).value
Pev = ws.cell(row =11 , column = 6).value
Peh = ws.cell(row =12 , column = 6).value
Pcv = ws.cell(row =13 , column = 6).value
Pch = ws.cell(row =14 , column = 6).value

theta = radians(30)
#Member forces computations
#Reactions
Rcv =  (Pdh*H/2 + Pdv*L/4 + Pfh*H + Pfv*L/2 + 3*Pev*L/4 + Peh * H/2) / L
Rav = Pav + Pcv + Pdv + Pfv + Pev - Rcv
Rah = Pah + Pdh + Pfh + Peh + Pch

# P forces
Fce = (Rcv - Pcv) / sin(theta)
Fch = Pch + Fce * cos(theta)
Fbh = Fch
Ffe = 0.5 * (Pev / sin(theta) + Peh / cos(theta))
Fbe = Ffe - Peh / cos(theta)
Ffd = -Pfh / cos(theta) - Ffe
Ffb = Pfv + Ffe*sin(theta) - Ffd*sin(theta)
Fdb = (-Ffb - Fbe * sin(theta)) / sin(theta)
Fgb = Fbh - Fbe * cos(theta) + Fdb * cos(theta)
Fag = Fgb
Fad = (Rav - Pav) / sin(theta)

#P1 Calculations
P1ag = P1gb = P1bh = P1hc = 1


file_name = "Book1-banmara-jhol.xlsx"