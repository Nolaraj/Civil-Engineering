from math import *

import openpyxl
import os

file_name = "Book1-banmara-jhol.xlsx"
Sheet_name = "Stress Strain"

folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads')
File_Path = os.path.join(folder, file_name)
Excel_file = os.path.join(folder, file_name)

wb = openpyxl.load_workbook(File_Path)
ws = wb[Sheet_name]


max_row= ws.max_row
Columns = [2, 8, 14, 20, 26]
start_row = 4


def Data_adjuster(Column):
    i = start_row
    while i <= max_row + 1:
        Del_L = ws.cell(row =i , column = Column).value
        if ws.cell(row =i + 1 , column = Column).value is not None :
            n = 1
            for j in range(i + 1, max_row + 1):
                if Del_L == ws.cell(row=j, column=Column).value:
                    n += 1

                elif ws.cell(row=j, column=Column).value is None:
                    Del_Ln = Del_L  + 1
                    break

                else:
                    Del_Ln  = ws.cell(row=j, column=Column).value
                    print(Del_Ln, j)
                    break
            for k in range(i, i + n):
                Del_L1 = Del_L + ((Del_Ln - Del_L) * (k - i)/n)
                ws.cell(row =k , column = Column).value = Del_L1

            i = i + n

        else:
            break

for Column in Columns:
    Data_adjuster(Column)



wb.save(File_Path)




